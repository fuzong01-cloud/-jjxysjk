from app.core.database import SessionLocal
from app.models.entities import Student
from openpyxl import Workbook


def test_health(client):
    assert client.get("/health").json() == {"status": "ok"}


def test_api_requires_login(client):
    assert client.get("/api/v1/students").status_code == 401


def test_search_update_and_audit(authenticated_client):
    with SessionLocal() as db:
        student = Student(name="测试学员", id_card_number="11010519491231002X", district_county="北京市延庆区")
        db.add(student); db.commit(); student_id = student.id
    result = authenticated_client.get("/api/v1/students", params={"name": "测试"})
    assert result.status_code == 200
    assert result.json()["total"] == 1
    update = authenticated_client.patch(f"/api/v1/students/{student_id}/basic-info", json={"phone": "13800138000"})
    assert update.status_code == 200
    logs = authenticated_client.get("/api/v1/audit-logs").json()["data"]
    assert any(row["action"] == "STUDENT_UPDATE" for row in logs)


def test_add_honor(authenticated_client):
    with SessionLocal() as db:
        student = Student(name="荣誉学员", id_card_number="11010519491231002X")
        db.add(student); db.commit(); student_id = student.id
    result = authenticated_client.post(f"/api/v1/students/{student_id}/honors", json={"honor_level": "市级", "honor_description": "优秀学员"})
    assert result.status_code == 200


def test_detail_module_create_update_delete(authenticated_client):
    with SessionLocal() as db:
        student = Student(name="模块维护学员", id_card_number="11010519491231002X")
        db.add(student); db.commit(); student_id = student.id

    education = authenticated_client.put(f"/api/v1/students/{student_id}/education", json={"education_level": "本科", "graduate_school": "测试大学", "major": "农学", "certificate_number": "A001", "learning_experience": "学习", "work_experience": "工作", "training_experience": "培训"})
    assert education.status_code == 200
    education2 = authenticated_client.put(f"/api/v1/students/{student_id}/education", json={"education_level": "研究生", "graduate_school": "测试大学", "major": "农学", "certificate_number": "A002", "learning_experience": "学习2", "work_experience": "工作2", "training_experience": "培训2"})
    assert education2.status_code == 200

    honor = authenticated_client.post(f"/api/v1/students/{student_id}/honors", json={"honor_number": "H1", "honor_time": "2024-01-01", "honor_level": "区级", "honor_description": "测试荣誉"})
    honor_id = honor.json()["data"]["id"]
    assert authenticated_client.patch(f"/api/v1/honors/{honor_id}", json={"honor_number": "H2", "honor_time": "2024-02-01", "honor_level": "市级", "honor_description": "修改荣誉"}).status_code == 200

    entity_payload = {"entity_name": "测试主体", "entity_intro": "简介", "entity_type": "合作社", "entity_subtype": "种植", "entity_industry_type": "蔬菜", "established_date": "2020-01-01", "registered_address": "地址", "unified_social_credit_code": "CODE1", "industry_years": 4, "student_position_in_entity": "负责人", "patent_applications": "无", "farmer_households_driven": 10, "technical_partner_count": 2, "technical_partners": "单位", "quality_inspection_org": "是", "quality_system_certification": "认证", "green_organic_geo_certification": "绿色", "entity_honors": "荣誉", "supporting_policies": "支持"}
    entity = authenticated_client.post(f"/api/v1/students/{student_id}/business-entities", json=entity_payload)
    entity_id = entity.json()["data"]["id"]
    entity_payload["entity_name"] = "测试主体改"
    assert authenticated_client.patch(f"/api/v1/business-entities/{entity_id}", json=entity_payload).status_code == 200

    revenue_payload = {"year": 2024, "operating_revenue": 100, "net_profit": 10, "fixed_asset_net_value": 20, "total_assets": 200, "total_liabilities": 50, "employee_count": 8, "current_assets": 80, "management_expense": 5, "government_subsidy_amount": 3}
    revenue = authenticated_client.post(f"/api/v1/business-entities/{entity_id}/revenue", json=revenue_payload)
    revenue_id = revenue.json()["data"]["id"]
    revenue_payload["operating_revenue"] = 120
    assert authenticated_client.patch(f"/api/v1/revenue/{revenue_id}", json=revenue_payload).status_code == 200
    assert authenticated_client.delete(f"/api/v1/revenue/{revenue_id}").status_code == 400

    industry = authenticated_client.post(f"/api/v1/business-entities/{entity_id}/industries", json={"industry_name": "蔬菜", "three_year_total_income": 300, "operation_years": 5})
    industry_id = industry.json()["data"]["id"]
    assert authenticated_client.patch(f"/api/v1/industries/{industry_id}", json={"industry_name": "水果", "three_year_total_income": 400, "operation_years": 6}).status_code == 200

    cultivation = authenticated_client.post(f"/api/v1/students/{student_id}/cultivations", json={"cultivation_year": 2024, "cultivation_needs": "需求", "training_experience": "经历"})
    cultivation_id = cultivation.json()["data"]["id"]
    assert authenticated_client.patch(f"/api/v1/cultivations/{cultivation_id}", json={"cultivation_year": 2025, "cultivation_needs": "需求2", "training_experience": "经历2"}).status_code == 200

    detail = authenticated_client.get(f"/students/{student_id}")
    assert "测试主体改" in detail.text and "修改" in detail.text and "删除" in detail.text and "添加培育信息" in detail.text

    assert authenticated_client.delete(f"/api/v1/industries/{industry_id}").status_code == 200
    assert authenticated_client.delete(f"/api/v1/cultivations/{cultivation_id}").status_code == 200
    assert authenticated_client.delete(f"/api/v1/honors/{honor_id}").status_code == 200
    assert authenticated_client.delete(f"/api/v1/business-entities/{entity_id}").status_code == 200
    assert authenticated_client.delete(f"/api/v1/students/{student_id}/education").status_code == 400


def test_change_password(authenticated_client):
    result = authenticated_client.post("/api/v1/auth/password", json={"current_password": "TestPass123!", "new_password": "NewTestPass456!"})
    assert result.status_code == 200
    authenticated_client.post("/logout")
    assert authenticated_client.post("/login", data={"username": "admin", "password": "NewTestPass456!"}).status_code == 200


def test_id_card_page_search_and_category_tables(authenticated_client):
    with SessionLocal() as db:
        student = Student(name="分类测试", id_card_number="11010519491231002X")
        db.add(student); db.commit(); student_id = student.id
    result = authenticated_client.get("/students", params={"id_card": "11010519491231002x"})
    assert result.status_code == 200 and "分类测试" in result.text and "身份证号（精准查询）" in result.text
    detail = authenticated_client.get(f"/students/{student_id}")
    for title in ("基本信息表", "受教育情况表", "荣誉情况表", "新型经营主体情况表", "近三年营收情况表", "主营产业情况表", "个人培育信息表"):
        assert title in detail.text


def test_delete_student_from_list(authenticated_client):
    with SessionLocal() as db:
        student = Student(name="待删除学员", id_card_number="11010519491231002X", district_county="测试区")
        db.add(student); db.commit(); student_id = student.id

    page = authenticated_client.get("/students", params={"name": "待删除"})
    assert page.status_code == 200 and "删除" in page.text and "批量处理" in page.text and "待删除学员" in page.text

    result = authenticated_client.delete(f"/api/v1/students/{student_id}")
    assert result.status_code == 200
    assert authenticated_client.get(f"/api/v1/students/{student_id}").status_code == 404
    assert authenticated_client.get("/students", params={"name": "待删除"}).text.count("待删除学员") == 0

    logs = authenticated_client.get("/api/v1/audit-logs").json()["data"]
    assert any(row["action"] == "STUDENT_DELETE" for row in logs)


def test_bulk_delete_students(authenticated_client):
    with SessionLocal() as db:
        one = Student(name="批量删除甲", id_card_number="11010519491231002X")
        two = Student(name="批量删除乙", id_card_number="110105194912310018")
        db.add_all([one, two]); db.commit(); ids = [one.id, two.id]

    result = authenticated_client.post("/api/v1/students/bulk-delete", json={"student_ids": ids})
    assert result.status_code == 200
    assert result.json()["data"]["deleted_count"] == 2
    assert authenticated_client.get("/api/v1/students", params={"name": "批量删除"}).json()["total"] == 0
    logs = authenticated_client.get("/api/v1/audit-logs").json()["data"]
    assert any(row["action"] == "STUDENT_BULK_DELETE" for row in logs)


def test_bulk_export_selected_students(authenticated_client):
    with SessionLocal() as db:
        one = Student(name="批量导出甲", id_card_number="11010519491231002X")
        two = Student(name="批量导出乙", id_card_number="110105194912310018")
        db.add_all([one, two]); db.commit(); ids = [one.id, two.id]

    page = authenticated_client.get("/students")
    assert "批量处理" in page.text and "已选择 0 位学员" not in page.text

    result = authenticated_client.get("/api/v1/export/students.xlsx", params={"ids": ",".join(map(str, ids))})
    assert result.status_code == 200
    assert result.headers["content-type"].startswith("application/vnd.openxmlformats")
    logs = authenticated_client.get("/api/v1/audit-logs").json()["data"]
    assert any(row["action"] == "EXPORT_SELECTED_STUDENTS" for row in logs)


def test_import_template_download(authenticated_client):
    result = authenticated_client.get("/api/v1/import/template.xlsx")
    assert result.status_code == 200
    assert result.headers["content-type"].startswith("application/vnd.openxmlformats")


def test_import_confirmation_preview_rows(authenticated_client, tmp_path):
    path = tmp_path / "preview.xlsx"
    book = Workbook(); sheet = book.active; sheet.title = "Sheet1"
    sheet.append(["学员姓名", "身份证号", "所在区县", "新型经营主体名称", "个人获得荣誉情况1"])
    sheet.append(["预览学员", "11010519491231002X", "北京市延庆区", "预览合作社", "优秀学员"])
    book.save(path)

    with path.open("rb") as handle:
        result = authenticated_client.post(
            "/api/v1/import/preview",
            files={"file": ("preview.xlsx", handle, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

    assert result.status_code == 200
    data = result.json()["data"]
    assert data["rows"][0]["name"] == "预览学员"
    assert data["rows"][0]["business_entity_name"] == "预览合作社"
    assert data["rows"][0]["modules"]["honors"] == 1
    assert data["rows"][0]["modules"]["business_entities"] is True


def test_missing_modules_warn_but_do_not_block_import(authenticated_client, tmp_path):
    path = tmp_path / "missing-modules.xlsx"
    book = Workbook(); sheet = book.active; sheet.title = "Sheet1"
    sheet.append(["学员姓名", "身份证号"])
    sheet.append(["缺模块学员", "11010519491231002X"])
    book.save(path)

    with path.open("rb") as handle:
        result = authenticated_client.post(
            "/api/v1/import/preview",
            files={"file": ("missing-modules.xlsx", handle, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

    data = result.json()["data"]
    assert data["failed_rows"] == 0
    assert data["rows"][0]["notice_messages"]
    commit = authenticated_client.post(f"/api/v1/import/batches/{data['batch_id']}/commit")
    assert commit.status_code == 200


def test_import_commit_rejects_failed_preview(authenticated_client, tmp_path):
    path = tmp_path / "bad-preview.xlsx"
    book = Workbook(); sheet = book.active; sheet.title = "Sheet1"
    sheet.append(["学员姓名", "身份证号"])
    sheet.append(["错误学员", ""])
    book.save(path)

    with path.open("rb") as handle:
        result = authenticated_client.post(
            "/api/v1/import/preview",
            files={"file": ("bad-preview.xlsx", handle, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

    batch_id = result.json()["data"]["batch_id"]
    error_row = [row for row in result.json()["data"]["rows"] if row["action"] == "错误"][0]
    assert error_row["has_error"] is True
    assert error_row["error_messages"]
    commit = authenticated_client.post(f"/api/v1/import/batches/{batch_id}/commit")
    assert commit.status_code == 400


def test_cancel_import_preview(authenticated_client, tmp_path):
    path = tmp_path / "cancel-preview.xlsx"
    book = Workbook(); sheet = book.active; sheet.title = "Sheet1"
    sheet.append(["学员姓名", "身份证号"])
    sheet.append(["取消预览学员", "11010519491231002X"])
    book.save(path)

    with path.open("rb") as handle:
        result = authenticated_client.post(
            "/api/v1/import/preview",
            files={"file": ("cancel-preview.xlsx", handle, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

    batch_id = result.json()["data"]["batch_id"]
    cancel = authenticated_client.delete(f"/api/v1/import/batches/{batch_id}")
    assert cancel.status_code == 200
    assert authenticated_client.get(f"/api/v1/import/batches/{batch_id}/preview").status_code == 404
