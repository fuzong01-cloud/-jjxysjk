import tempfile
from datetime import datetime
from pathlib import Path

from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, UploadFile
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import Workbook
from sqlalchemy import func, or_, select
from sqlalchemy.orm import Session, selectinload

from app.core.auth import current_user
from app.core.config import get_settings
from app.core.database import get_db
from app.core.security import hash_password, mask_id_card, mask_phone, verify_password
from app.models.entities import AnnualRevenueRecord, AuditLog, BackupRecord, BusinessEntity, CultivationRecord, EducationRecord, HonorRecord, ImportBatch, ImportError, MainIndustry, Student, User
from app.schemas.api import ApiResponse, CultivationCreate, EducationUpdate, EntityUpdate, HonorCreate, IndustryCreate, PasswordChange, RevenueCreate, StudentBulkDelete, StudentUpdate
from app.services.audit_service import add_audit
from app.services.backup_service import create_backup, restore_database
from app.services.import_service import commit_batch, create_preview, preview_rows_for_batch


router = APIRouter(prefix="/api/v1")


def student_query(name: str | None, id_card: str | None, district: str | None, entity_name: str | None):
    stmt = select(Student).where(Student.deleted_at.is_(None))
    if name: stmt = stmt.where(Student.name.contains(name))
    if id_card: stmt = stmt.where(Student.id_card_number == id_card.replace(" ", "").upper())
    if district: stmt = stmt.where(Student.district_county.contains(district))
    if entity_name: stmt = stmt.join(Student.entities).where(BusinessEntity.entity_name.contains(entity_name))
    return stmt


def archive_data(student: Student) -> dict:
    """返回七个需求模块的完整学员档案。"""
    entities = [entity for entity in student.entities if entity.deleted_at is None]
    return {
        "basic_info": {
            "id": student.id, "id_card_number": student.id_card_number, "name": student.name,
            "gender": student.gender, "birth_date": student.birth_date, "age": student.age,
            "ethnicity": student.ethnicity, "native_place": student.native_place,
            "district_county": student.district_county, "political_status": student.political_status,
            "phone": student.phone, "health_status": student.health_status,
            "professional_title": student.professional_title, "wechat": student.wechat,
            "email": student.email, "household_type": student.household_type,
            "postal_code": student.postal_code, "home_address": student.home_address,
            "talent_category": student.talent_category, "status": student.status,
            "administrative_position": student.administrative_position,
            "social_part_time_positions": student.social_part_time_positions,
        },
        "education": None if student.education is None else {
            "education_level": student.education.education_level,
            "graduate_school": student.education.graduate_school, "major": student.education.major,
            "certificate_number": student.education.certificate_number,
            "learning_experience": student.education.learning_experience,
            "work_experience": student.education.work_experience,
            "training_experience": student.education.training_experience,
        },
        "honors": [
            {"id": honor.id, "honor_time": honor.honor_time, "honor_level": honor.honor_level,
             "honor_description": honor.honor_description}
            for honor in student.honors if honor.deleted_at is None
        ],
        "business_entities": [
            {"id": entity.id, "entity_name": entity.entity_name, "entity_intro": entity.entity_intro,
             "entity_type": entity.entity_type, "entity_subtype": entity.entity_subtype,
             "established_date": entity.established_date, "registered_address": entity.registered_address,
             "unified_social_credit_code": entity.unified_social_credit_code,
             "student_position_in_entity": entity.student_position_in_entity}
            for entity in entities
        ],
        "annual_revenues": [
            {"id": revenue.id, "business_entity_id": entity.id, "entity_name": entity.entity_name,
             "year": revenue.year, "operating_revenue": revenue.operating_revenue,
             "net_profit": revenue.net_profit, "fixed_asset_net_value": revenue.fixed_asset_net_value,
             "total_assets": revenue.total_assets, "total_liabilities": revenue.total_liabilities,
             "employee_count": revenue.employee_count, "current_assets": revenue.current_assets,
             "management_expense": revenue.management_expense,
             "government_subsidy_amount": revenue.government_subsidy_amount}
            for entity in entities for revenue in entity.revenues
        ],
        "main_industries": [
            {"id": industry.id, "business_entity_id": entity.id, "entity_name": entity.entity_name,
             "industry_name": industry.industry_name,
             "three_year_total_income": industry.three_year_total_income,
             "operation_years": industry.operation_years}
            for entity in entities for industry in entity.industries
        ],
        "cultivations": [
            {"id": cultivation.id, "cultivation_year": cultivation.cultivation_year,
             "cultivation_needs": cultivation.cultivation_needs,
             "training_experience": cultivation.training_experience}
            for cultivation in student.cultivations
        ],
    }


@router.get("/auth/me")
def me(user: User = Depends(current_user)) -> ApiResponse:
    return ApiResponse(data={"id": user.id, "username": user.username})


@router.post("/auth/password")
def change_password(payload: PasswordChange, db: Session = Depends(get_db), user: User = Depends(current_user)) -> ApiResponse:
    if not verify_password(payload.current_password, user.password_hash):
        raise HTTPException(400, "当前密码不正确")
    if payload.current_password == payload.new_password:
        raise HTTPException(400, "新密码不能与当前密码相同")
    user.password_hash = hash_password(payload.new_password)
    add_audit(db, user_id=user.id, action="PASSWORD_CHANGE", target_table="users", target_id=user.id)
    db.commit()
    return ApiResponse(message="密码已修改")


@router.get("/students")
def students(name: str | None = None, id_card: str | None = None, district: str | None = None, entity_name: str | None = None, page: int = Query(1, ge=1), page_size: int = Query(20, ge=1, le=100), db: Session = Depends(get_db), _user: User = Depends(current_user)) -> dict:
    stmt = student_query(name, id_card, district, entity_name)
    total = db.scalar(select(func.count()).select_from(stmt.subquery())) or 0
    rows = db.scalars(stmt.order_by(Student.id.desc()).offset((page - 1) * page_size).limit(page_size)).all()
    return {"items": [{"id": s.id, "name": s.name, "id_card_number": mask_id_card(s.id_card_number), "phone": mask_phone(s.phone), "district_county": s.district_county, "status": s.status, "age": s.age} for s in rows], "total": total, "page": page, "page_size": page_size}


@router.get("/students/{student_id}")
def student_detail(student_id: int, db: Session = Depends(get_db), _user: User = Depends(current_user)) -> ApiResponse:
    student = db.scalar(select(Student).options(selectinload(Student.education), selectinload(Student.honors), selectinload(Student.entities).selectinload(BusinessEntity.revenues), selectinload(Student.entities).selectinload(BusinessEntity.industries), selectinload(Student.cultivations)).where(Student.id == student_id, Student.deleted_at.is_(None)))
    if not student: raise HTTPException(404, "学员不存在")
    return ApiResponse(data=archive_data(student))


@router.get("/students/by-id-card/{id_card_number}")
def by_id_card(id_card_number: str, db: Session = Depends(get_db), user: User = Depends(current_user)) -> ApiResponse:
    student = db.scalar(select(Student).where(Student.id_card_number == id_card_number.replace(" ", "").upper(), Student.deleted_at.is_(None)))
    if not student: raise HTTPException(404, "学员不存在")
    return student_detail(student.id, db, user)


@router.patch("/students/{student_id}/basic-info")
def update_student(student_id: int, payload: StudentUpdate, request: Request, db: Session = Depends(get_db), user: User = Depends(current_user)) -> ApiResponse:
    student = db.get(Student, student_id)
    if not student or student.deleted_at is not None: raise HTTPException(404, "学员不存在")
    changes = payload.model_dump(exclude_unset=True); before = {k: getattr(student, k) for k in changes}
    if "id_card_number" in changes and changes["id_card_number"]:
        new_card = changes["id_card_number"].replace(" ", "").upper()
        exists = db.scalar(select(Student).where(Student.id_card_number == new_card, Student.id != student.id))
        if exists:
            raise HTTPException(400, "身份证号已存在")
        changes["id_card_number"] = new_card
    for key, value in changes.items(): setattr(student, key, value)
    add_audit(db, user_id=user.id, action="STUDENT_UPDATE", target_table="students", target_id=student.id, before=before, after=changes, ip_address=request.client.host if request.client else None)
    db.commit(); return ApiResponse(data={"id": student.id}, message="学员信息已更新")


@router.delete("/students/{student_id}")
def delete_student(student_id: int, db: Session = Depends(get_db), user: User = Depends(current_user)) -> ApiResponse:
    student = db.get(Student, student_id)
    if not student or student.deleted_at is not None:
        raise HTTPException(404, "学员不存在")
    before = {"name": student.name, "id_card_number": mask_id_card(student.id_card_number)}
    student.deleted_at = datetime.utcnow()
    add_audit(db, user_id=user.id, action="STUDENT_DELETE", target_table="students", target_id=student.id, before=before)
    db.commit()
    return ApiResponse(data={"id": student.id}, message="学员档案已删除")


@router.post("/students/bulk-delete")
def bulk_delete_students(payload: StudentBulkDelete, db: Session = Depends(get_db), user: User = Depends(current_user)) -> ApiResponse:
    ids = list(dict.fromkeys(payload.student_ids))
    students = db.scalars(select(Student).where(Student.id.in_(ids), Student.deleted_at.is_(None))).all()
    if not students:
        raise HTTPException(404, "没有可删除的学员")
    now = datetime.utcnow()
    for student in students:
        student.deleted_at = now
    add_audit(
        db,
        user_id=user.id,
        action="STUDENT_BULK_DELETE",
        target_table="students",
        before={"count": len(students), "ids": [student.id for student in students]},
    )
    db.commit()
    return ApiResponse(data={"deleted_count": len(students)}, message="已批量删除")


@router.put("/students/{student_id}/education")
def upsert_education(student_id: int, payload: EducationUpdate, db: Session = Depends(get_db), user: User = Depends(current_user)) -> ApiResponse:
    student = db.scalar(select(Student).options(selectinload(Student.education)).where(Student.id == student_id, Student.deleted_at.is_(None)))
    if not student:
        raise HTTPException(404, "学员不存在")
    data = payload.model_dump()
    if student.education is None:
        student.education = EducationRecord(**data)
        db.flush()
        target_id = student.education.id
        action = "EDUCATION_CREATE"
        before = None
    else:
        before = {k: getattr(student.education, k) for k in data}
        for key, value in data.items():
            setattr(student.education, key, value)
        target_id = student.education.id
        action = "EDUCATION_UPDATE"
    add_audit(db, user_id=user.id, action=action, target_table="education_records", target_id=target_id, before=before, after=data)
    db.commit()
    return ApiResponse(data={"id": target_id}, message="受教育情况已保存")


@router.delete("/students/{student_id}/education")
def delete_education(student_id: int, db: Session = Depends(get_db), user: User = Depends(current_user)) -> ApiResponse:
    student = db.scalar(select(Student).options(selectinload(Student.education)).where(Student.id == student_id, Student.deleted_at.is_(None)))
    if not student or student.education is None:
        raise HTTPException(404, "受教育情况不存在")
    target_id = student.education.id
    db.delete(student.education)
    add_audit(db, user_id=user.id, action="EDUCATION_DELETE", target_table="education_records", target_id=target_id)
    db.commit()
    return ApiResponse(message="受教育情况已删除")


@router.post("/students/{student_id}/honors")
def add_honor(student_id: int, payload: HonorCreate, db: Session = Depends(get_db), user: User = Depends(current_user)) -> ApiResponse:
    if not db.get(Student, student_id): raise HTTPException(404, "学员不存在")
    honor = HonorRecord(student_id=student_id, **payload.model_dump()); db.add(honor); db.flush()
    add_audit(db, user_id=user.id, action="HONOR_CREATE", target_table="honor_records", target_id=honor.id, after=payload.model_dump(mode="json")); db.commit()
    return ApiResponse(data={"id": honor.id}, message="荣誉已添加")


@router.patch("/honors/{honor_id}")
def update_honor(honor_id: int, payload: HonorCreate, db: Session = Depends(get_db), user: User = Depends(current_user)) -> ApiResponse:
    honor = db.get(HonorRecord, honor_id)
    if not honor or honor.deleted_at is not None:
        raise HTTPException(404, "荣誉不存在")
    changes = payload.model_dump(); before = {k: getattr(honor, k) for k in changes}
    for key, value in changes.items():
        setattr(honor, key, value)
    add_audit(db, user_id=user.id, action="HONOR_UPDATE", target_table="honor_records", target_id=honor.id, before=before, after=changes)
    db.commit()
    return ApiResponse(data={"id": honor.id}, message="荣誉已更新")


@router.delete("/honors/{honor_id}")
def delete_honor(honor_id: int, db: Session = Depends(get_db), user: User = Depends(current_user)) -> ApiResponse:
    honor = db.get(HonorRecord, honor_id)
    if not honor: raise HTTPException(404, "荣誉不存在")
    from datetime import datetime
    honor.deleted_at = datetime.utcnow(); add_audit(db, user_id=user.id, action="HONOR_DELETE", target_table="honor_records", target_id=honor.id); db.commit()
    return ApiResponse(message="荣誉已删除")


@router.patch("/business-entities/{entity_id}")
def update_entity(entity_id: int, payload: EntityUpdate, db: Session = Depends(get_db), user: User = Depends(current_user)) -> ApiResponse:
    entity = db.get(BusinessEntity, entity_id)
    if not entity: raise HTTPException(404, "经营主体不存在")
    changes = payload.model_dump(exclude_unset=True); before = {k: getattr(entity, k) for k in changes}
    for key, value in changes.items(): setattr(entity, key, value)
    add_audit(db, user_id=user.id, action="ENTITY_UPDATE", target_table="business_entities", target_id=entity.id, before=before, after=changes); db.commit()
    return ApiResponse(data={"id": entity.id}, message="经营主体已更新")


@router.post("/students/{student_id}/business-entities")
def add_entity(student_id: int, payload: EntityUpdate, db: Session = Depends(get_db), user: User = Depends(current_user)) -> ApiResponse:
    if not db.scalar(select(Student.id).where(Student.id == student_id, Student.deleted_at.is_(None))):
        raise HTTPException(404, "学员不存在")
    data = payload.model_dump()
    entity = BusinessEntity(student_id=student_id, **data)
    db.add(entity); db.flush()
    add_audit(db, user_id=user.id, action="ENTITY_CREATE", target_table="business_entities", target_id=entity.id, after=data)
    db.commit()
    return ApiResponse(data={"id": entity.id}, message="经营主体已添加")


@router.delete("/business-entities/{entity_id}")
def delete_entity(entity_id: int, db: Session = Depends(get_db), user: User = Depends(current_user)) -> ApiResponse:
    entity = db.get(BusinessEntity, entity_id)
    if not entity or entity.deleted_at is not None:
        raise HTTPException(404, "经营主体不存在")
    entity.deleted_at = datetime.utcnow()
    add_audit(db, user_id=user.id, action="ENTITY_DELETE", target_table="business_entities", target_id=entity.id)
    db.commit()
    return ApiResponse(message="经营主体已删除")


@router.get("/business-entities")
def entity_list(entity_name: str | None = None, student_name: str | None = None, district: str | None = None, page: int = Query(1, ge=1), page_size: int = Query(20, ge=1, le=100), db: Session = Depends(get_db), _user: User = Depends(current_user)) -> dict:
    stmt = select(BusinessEntity).join(BusinessEntity.student).where(BusinessEntity.deleted_at.is_(None))
    if entity_name: stmt = stmt.where(BusinessEntity.entity_name.contains(entity_name))
    if student_name: stmt = stmt.where(Student.name.contains(student_name))
    if district: stmt = stmt.where(Student.district_county.contains(district))
    total = db.scalar(select(func.count()).select_from(stmt.subquery())) or 0
    rows = db.scalars(stmt.offset((page - 1) * page_size).limit(page_size)).all()
    return {"items": [{"id": e.id, "entity_name": e.entity_name, "entity_type": e.entity_type, "student_id": e.student_id, "student_name": e.student.name} for e in rows], "total": total, "page": page, "page_size": page_size}


@router.get("/business-entities/{entity_id}")
def entity_detail(entity_id: int, db: Session = Depends(get_db), _user: User = Depends(current_user)) -> ApiResponse:
    entity = db.scalar(select(BusinessEntity).options(selectinload(BusinessEntity.revenues), selectinload(BusinessEntity.industries)).where(BusinessEntity.id == entity_id))
    if not entity: raise HTTPException(404, "经营主体不存在")
    return ApiResponse(data={
        "id": entity.id, "student_id": entity.student_id, "entity_name": entity.entity_name,
        "entity_intro": entity.entity_intro, "entity_type": entity.entity_type,
        "entity_subtype": entity.entity_subtype, "entity_industry_type": entity.entity_industry_type,
        "established_date": entity.established_date, "registered_address": entity.registered_address,
        "unified_social_credit_code": entity.unified_social_credit_code,
        "industry_years": entity.industry_years,
        "student_position_in_entity": entity.student_position_in_entity,
        "patent_applications": entity.patent_applications,
        "farmer_households_driven": entity.farmer_households_driven,
        "technical_partner_count": entity.technical_partner_count,
        "technical_partners": entity.technical_partners,
        "quality_inspection_org": entity.quality_inspection_org,
        "quality_system_certification": entity.quality_system_certification,
        "green_organic_geo_certification": entity.green_organic_geo_certification,
        "entity_honors": entity.entity_honors,
        "supporting_policies": entity.supporting_policies,
    })


@router.post("/business-entities/{entity_id}/revenue")
def add_revenue(entity_id: int, payload: RevenueCreate, db: Session = Depends(get_db), user: User = Depends(current_user)) -> ApiResponse:
    if not db.get(BusinessEntity, entity_id): raise HTTPException(404, "经营主体不存在")
    record = AnnualRevenueRecord(business_entity_id=entity_id, **payload.model_dump()); db.add(record); db.flush(); add_audit(db, user_id=user.id, action="REVENUE_CREATE", target_table="annual_revenue_records", target_id=record.id, after=payload.model_dump()); db.commit()
    return ApiResponse(data={"id": record.id}, message="年度营收已添加")


@router.get("/business-entities/{entity_id}/revenue")
def revenues(entity_id: int, db: Session = Depends(get_db), _user: User = Depends(current_user)) -> ApiResponse:
    rows = db.scalars(select(AnnualRevenueRecord).where(AnnualRevenueRecord.business_entity_id == entity_id).order_by(AnnualRevenueRecord.year.desc())).all()
    return ApiResponse(data=[{
        "id": r.id, "year": r.year, "operating_revenue": r.operating_revenue,
        "net_profit": r.net_profit, "fixed_asset_net_value": r.fixed_asset_net_value,
        "total_assets": r.total_assets, "total_liabilities": r.total_liabilities,
        "employee_count": r.employee_count, "current_assets": r.current_assets,
        "management_expense": r.management_expense,
        "government_subsidy_amount": r.government_subsidy_amount,
    } for r in rows])


@router.patch("/revenue/{revenue_id}")
def update_revenue(revenue_id: int, payload: RevenueCreate, db: Session = Depends(get_db), user: User = Depends(current_user)) -> ApiResponse:
    record = db.get(AnnualRevenueRecord, revenue_id)
    if not record: raise HTTPException(404, "年度营收不存在")
    changes = payload.model_dump(exclude_unset=True); before = {k: getattr(record, k) for k in changes}
    for key, value in changes.items(): setattr(record, key, value)
    add_audit(db, user_id=user.id, action="REVENUE_UPDATE", target_table="annual_revenue_records", target_id=record.id, before=before, after=changes); db.commit()
    return ApiResponse(data={"id": record.id}, message="年度营收已更新")


@router.delete("/revenue/{revenue_id}")
def delete_revenue(revenue_id: int, db: Session = Depends(get_db), user: User = Depends(current_user)) -> ApiResponse:
    record = db.get(AnnualRevenueRecord, revenue_id)
    if not record:
        raise HTTPException(404, "年度营收不存在")
    db.delete(record)
    add_audit(db, user_id=user.id, action="REVENUE_DELETE", target_table="annual_revenue_records", target_id=revenue_id)
    db.commit()
    return ApiResponse(message="年度营收已删除")


@router.post("/business-entities/{entity_id}/industries")
def add_industry(entity_id: int, payload: IndustryCreate, db: Session = Depends(get_db), user: User = Depends(current_user)) -> ApiResponse:
    if not db.get(BusinessEntity, entity_id): raise HTTPException(404, "经营主体不存在")
    record = MainIndustry(business_entity_id=entity_id, **payload.model_dump()); db.add(record); db.flush(); add_audit(db, user_id=user.id, action="INDUSTRY_CREATE", target_table="main_industries", target_id=record.id, after=payload.model_dump()); db.commit()
    return ApiResponse(data={"id": record.id}, message="主营产业已添加")


@router.get("/business-entities/{entity_id}/industries")
def industries(entity_id: int, db: Session = Depends(get_db), _user: User = Depends(current_user)) -> ApiResponse:
    rows = db.scalars(select(MainIndustry).where(MainIndustry.business_entity_id == entity_id)).all()
    return ApiResponse(data=[{"id": r.id, "industry_name": r.industry_name, "three_year_total_income": r.three_year_total_income, "operation_years": r.operation_years} for r in rows])


@router.patch("/industries/{industry_id}")
def update_industry(industry_id: int, payload: IndustryCreate, db: Session = Depends(get_db), user: User = Depends(current_user)) -> ApiResponse:
    record = db.get(MainIndustry, industry_id)
    if not record: raise HTTPException(404, "主营产业不存在")
    changes = payload.model_dump(exclude_unset=True); before = {k: getattr(record, k) for k in changes}
    for key, value in changes.items(): setattr(record, key, value)
    add_audit(db, user_id=user.id, action="INDUSTRY_UPDATE", target_table="main_industries", target_id=record.id, before=before, after=changes); db.commit()
    return ApiResponse(data={"id": record.id}, message="主营产业已更新")


@router.delete("/industries/{industry_id}")
def delete_industry(industry_id: int, db: Session = Depends(get_db), user: User = Depends(current_user)) -> ApiResponse:
    record = db.get(MainIndustry, industry_id)
    if not record:
        raise HTTPException(404, "主营产业不存在")
    db.delete(record)
    add_audit(db, user_id=user.id, action="INDUSTRY_DELETE", target_table="main_industries", target_id=industry_id)
    db.commit()
    return ApiResponse(message="主营产业已删除")


@router.post("/students/{student_id}/cultivations")
def add_cultivation(student_id: int, payload: CultivationCreate, db: Session = Depends(get_db), user: User = Depends(current_user)) -> ApiResponse:
    if not db.scalar(select(Student.id).where(Student.id == student_id, Student.deleted_at.is_(None))):
        raise HTTPException(404, "学员不存在")
    data = payload.model_dump()
    record = CultivationRecord(student_id=student_id, **data)
    db.add(record); db.flush()
    add_audit(db, user_id=user.id, action="CULTIVATION_CREATE", target_table="cultivation_records", target_id=record.id, after=data)
    db.commit()
    return ApiResponse(data={"id": record.id}, message="培育信息已添加")


@router.patch("/cultivations/{cultivation_id}")
def update_cultivation(cultivation_id: int, payload: CultivationCreate, db: Session = Depends(get_db), user: User = Depends(current_user)) -> ApiResponse:
    record = db.get(CultivationRecord, cultivation_id)
    if not record:
        raise HTTPException(404, "培育信息不存在")
    changes = payload.model_dump(); before = {k: getattr(record, k) for k in changes}
    for key, value in changes.items():
        setattr(record, key, value)
    add_audit(db, user_id=user.id, action="CULTIVATION_UPDATE", target_table="cultivation_records", target_id=record.id, before=before, after=changes)
    db.commit()
    return ApiResponse(data={"id": record.id}, message="培育信息已更新")


@router.delete("/cultivations/{cultivation_id}")
def delete_cultivation(cultivation_id: int, db: Session = Depends(get_db), user: User = Depends(current_user)) -> ApiResponse:
    record = db.get(CultivationRecord, cultivation_id)
    if not record:
        raise HTTPException(404, "培育信息不存在")
    db.delete(record)
    add_audit(db, user_id=user.id, action="CULTIVATION_DELETE", target_table="cultivation_records", target_id=cultivation_id)
    db.commit()
    return ApiResponse(message="培育信息已删除")


@router.post("/import/preview")
async def import_preview(file: UploadFile = File(...), db: Session = Depends(get_db), user: User = Depends(current_user)) -> ApiResponse:
    if not file.filename or not file.filename.lower().endswith(".xlsx"): raise HTTPException(400, "仅支持 .xlsx 文件")
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp:
        temp.write(await file.read()); path = Path(temp.name)
    try: batch = create_preview(db, path, file.filename, user.id, get_settings().upload_dir)
    finally: path.unlink(missing_ok=True)
    return ApiResponse(
        data={
            "batch_id": batch.id,
            "total_rows": batch.total_rows,
            "success_rows": batch.success_rows,
            "new_rows": batch.new_rows,
            "updated_rows": batch.updated_rows,
            "failed_rows": batch.failed_rows,
            "warning_count": batch.warning_count,
            "preview_limit": 100,
            "rows": preview_rows_for_batch(db, batch, limit=100),
        },
        message="预览完成",
    )


@router.get("/import/template.xlsx")
def import_template(_user: User = Depends(current_user)):
    headers = [
        "学员姓名", "身份证号", "状态", "培育年份", "所在区县", "手机号", "性别", "出生年月日", "民族", "籍贯", "年龄",
        "政治面貌", "人才类别", "健康状况", "职称", "微信号", "邮箱", "户口性质", "邮编", "家庭住址", "行政职务", "社会兼职",
        "文化程度", "毕业院校", "所学专业", "证书编号", "学习经历", "工作经历", "培训经历",
        "荣誉编号1", "个人获得荣誉时间1", "个人获得荣誉级别1", "个人获得荣誉情况1",
        "荣誉编号2", "个人获得荣誉时间2", "个人获得荣誉级别2", "个人获得荣誉情况2",
        "新型经营主体名称", "新型经营主体简介", "新型经营主体类型", "新型经营主体细分类型", "主体产业类型", "成立日期", "登记住址",
        "统一社会信用代码", "从事相关产业年限", "在新型主体中职务", "近三年专利申请情况", "近三年带动农民（户）数量",
        "相关技术合作单位（家）", "合作单位分别为", "是否建有专门质检机构", "是否通过ISO9000、HACCP、GAP、GMP等质量体系认证",
        "是否获得“绿色食品、有机农产品和农产品地理标志”认证", "新型经营主体获得重要奖励及荣誉情况", "新型经营主体已获得发展配套支持",
    ]
    for index in range(1, 4):
        headers.extend([f"营收年份{index}", f"营业收入（万元）{index}", f"净利润（万元）{index}", f"固定资产净值（万元）{index}", f"总资产（万元）{index}", f"负债总额（万元）{index}", f"从业人数（人）{index}", f"流动资产（万元）{index}", f"管理费用（万元）{index}", f"政府补贴金额（万元）{index}"])
    for index in range(1, 4):
        headers.extend([f"主营产业{index}", f"近三年经营总收入（万元）{index}", f"经营年限{index}"])
    headers.append("培育诉求")
    book = Workbook(); sheet = book.active; sheet.title = "Sheet1"; sheet.append(headers); sheet.freeze_panes = "A2"; sheet.auto_filter.ref = f"A1:{sheet.cell(1, len(headers)).coordinate}"
    note = book.create_sheet("填写说明"); note.append(["项目", "说明"]); note.append(["必填字段", "学员姓名、身份证号"]); note.append(["数据结构", "第一行为字段名；从第二行开始每行一位学员"]); note.append(["重复信息", "荣誉、年度营收、主营产业使用带序号的列组填写"]); note.append(["日期", "建议使用 YYYY-MM-DD"]); note.append(["金额", "填写数字，单位万元"])
    stream = tempfile.SpooledTemporaryFile(); book.save(stream); stream.seek(0)
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=touyan_import_template.xlsx"})


@router.post("/import/batches/{batch_id}/commit")
def import_commit(batch_id: int, db: Session = Depends(get_db), user: User = Depends(current_user)) -> ApiResponse:
    batch = db.get(ImportBatch, batch_id)
    if not batch: raise HTTPException(404, "导入批次不存在")
    if batch.failed_rows:
        raise HTTPException(400, "预览存在错误行，请修正 Excel 后重新上传")
    try: commit_batch(db, batch, user.id)
    except ValueError as exc: raise HTTPException(400, str(exc)) from exc
    return ApiResponse(data={"batch_id": batch.id}, message="导入完成")


@router.get("/import/batches")
def batches(db: Session = Depends(get_db), _user: User = Depends(current_user)) -> ApiResponse:
    rows = db.scalars(select(ImportBatch).order_by(ImportBatch.id.desc()).limit(100)).all()
    return ApiResponse(data=[{"id": r.id, "file_name": r.file_name, "status": r.status, "total_rows": r.total_rows, "success_rows": r.success_rows, "failed_rows": r.failed_rows, "started_at": r.started_at} for r in rows])


@router.get("/import/batches/{batch_id}/preview")
def batch_preview(batch_id: int, db: Session = Depends(get_db), _user: User = Depends(current_user)) -> ApiResponse:
    batch = db.get(ImportBatch, batch_id)
    if not batch:
        raise HTTPException(404, "导入批次不存在")
    return ApiResponse(
        data={
            "batch_id": batch.id,
            "file_name": batch.file_name,
            "status": batch.status,
            "total_rows": batch.total_rows,
            "success_rows": batch.success_rows,
            "new_rows": batch.new_rows,
            "updated_rows": batch.updated_rows,
            "failed_rows": batch.failed_rows,
            "warning_count": batch.warning_count,
            "preview_limit": 100,
            "rows": preview_rows_for_batch(db, batch, limit=100),
        }
    )


@router.delete("/import/batches/{batch_id}")
def cancel_import_batch(batch_id: int, db: Session = Depends(get_db), _user: User = Depends(current_user)) -> ApiResponse:
    batch = db.get(ImportBatch, batch_id)
    if not batch:
        raise HTTPException(404, "导入批次不存在")
    if batch.status != "preview":
        raise HTTPException(400, "只能取消尚未写入的预览批次")
    db.delete(batch)
    db.commit()
    return ApiResponse(data={"batch_id": batch_id}, message="已取消本次预览")


@router.get("/import/batches/{batch_id}/errors")
def batch_errors(batch_id: int, db: Session = Depends(get_db), _user: User = Depends(current_user)) -> ApiResponse:
    rows = db.scalars(select(ImportError).where(ImportError.import_batch_id == batch_id)).all()
    return ApiResponse(data=[{"row": r.excel_row_number, "field": r.field_name, "message": r.error_message, "severity": r.severity} for r in rows])


@router.get("/export/students.xlsx")
def export_students(name: str | None = None, district: str | None = None, ids: str | None = None, db: Session = Depends(get_db), user: User = Depends(current_user)):
    if ids:
        try:
            selected_ids = [int(item) for item in ids.split(",") if item.strip()]
        except ValueError as exc:
            raise HTTPException(400, "学员 ID 格式不正确") from exc
        rows = db.scalars(select(Student).where(Student.id.in_(selected_ids), Student.deleted_at.is_(None)).order_by(Student.id)).all()
        export_action = "EXPORT_SELECTED_STUDENTS"
    else:
        rows = db.scalars(student_query(name, None, district, None).order_by(Student.id)).all()
        export_action = "EXPORT_STUDENTS"
    book = Workbook(); sheet = book.active; sheet.title = "学员名单"; sheet.append(["姓名", "身份证号", "性别", "出生日期", "年龄", "所在区县", "手机号", "状态"])
    for s in rows: sheet.append([s.name, s.id_card_number, s.gender, s.birth_date, s.age, s.district_county, s.phone, s.status])
    stream = tempfile.SpooledTemporaryFile(); book.save(stream); stream.seek(0)
    add_audit(db, user_id=user.id, action=export_action, target_table="students", after={"count": len(rows)}); db.commit()
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=students.xlsx"})


@router.get("/audit-logs")
def audit_logs(action: str | None = None, db: Session = Depends(get_db), _user: User = Depends(current_user)) -> ApiResponse:
    stmt = select(AuditLog).order_by(AuditLog.id.desc()).limit(200)
    if action: stmt = stmt.where(AuditLog.action == action)
    rows = db.scalars(stmt).all(); return ApiResponse(data=[{"id": r.id, "action": r.action, "target_table": r.target_table, "target_id": r.target_id, "created_at": r.created_at, "before": r.before_data, "after": r.after_data} for r in rows])


@router.post("/backups")
def backup(db: Session = Depends(get_db), user: User = Depends(current_user)) -> ApiResponse:
    record = create_backup(db, user.id); add_audit(db, user_id=user.id, action="BACKUP_CREATE", target_table="backup_records", target_id=record.id); db.commit()
    return ApiResponse(data={"id": record.id, "file_size": record.file_size}, message="备份已创建")


@router.get("/backups")
def backups(db: Session = Depends(get_db), _user: User = Depends(current_user)) -> ApiResponse:
    rows = db.scalars(select(BackupRecord).order_by(BackupRecord.id.desc())).all(); return ApiResponse(data=[{"id": r.id, "file": Path(r.backup_file_path).name, "size": r.file_size, "type": r.backup_type, "created_at": r.created_at} for r in rows])


@router.get("/backups/{backup_id}/download")
def download_backup(backup_id: int, db: Session = Depends(get_db), _user: User = Depends(current_user)):
    record = db.get(BackupRecord, backup_id)
    if not record or not Path(record.backup_file_path).exists(): raise HTTPException(404, "备份不存在")
    return FileResponse(record.backup_file_path, filename=Path(record.backup_file_path).name)


@router.post("/backups/restore")
async def restore_backup(file: UploadFile = File(...), db: Session = Depends(get_db), user: User = Depends(current_user)) -> ApiResponse:
    with tempfile.NamedTemporaryFile(delete=False, suffix=".db") as temp:
        temp.write(await file.read()); path = Path(temp.name)
    try:
        restore_database(db, path, user.id)
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    finally:
        path.unlink(missing_ok=True)
    return ApiResponse(message="数据库已恢复，请重新登录")
