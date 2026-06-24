import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.core.validators import clean_text


BASIC_FIELDS = {
    "学员姓名": "name", "姓名": "name", "状态": "status", "身份证号": "id_card_number",
    "性别": "gender", "出生年月日": "birth_date", "民族": "ethnicity", "籍贯": "native_place",
    "年龄": "age_snapshot", "政治面貌": "political_status", "所在区县": "district_county",
    "手机号": "phone", "健康状况": "health_status", "职称": "professional_title", "微信号": "wechat",
    "邮箱": "email", "户口性质": "household_type", "邮编": "postal_code", "家庭住址": "home_address",
    "人才类别": "talent_category", "行政职务": "administrative_position", "社会兼职": "social_part_time_positions",
}
EDUCATION_FIELDS = {
    "文化程度": "education_level", "毕业院校": "graduate_school", "所学专业": "major",
    "证书编号": "certificate_number", "学习经历": "learning_experience", "工作经历": "work_experience",
    "培训经历": "training_experience",
}
ENTITY_FIELDS = {
    "新型经营主体名称": "entity_name", "新型经营主体简介": "entity_intro", "新型经营主体类型": "entity_type",
    "新型经营主体细分类型": "entity_subtype", "成立日期": "established_date", "登记住址": "registered_address",
    "主体产业类型": "entity_industry_type",
    "统一社会信用代码": "unified_social_credit_code", "从事相关产业年限": "industry_years",
    "在新型主体中职务": "student_position_in_entity", "近三年专利申请情况": "patent_applications",
    "近三年带动农民（户）数量": "farmer_households_driven", "相关技术合作单位（家）": "technical_partner_count",
    "合作单位分别为": "technical_partners", "是否建有专门质检机构": "quality_inspection_org",
    "是否通过ISO9000、HACCP、GAP、GMP等质量体系认证": "quality_system_certification",
    "是否获得“绿色食品、有机农产品和农产品地理标志”认证": "green_organic_geo_certification",
    "新型经营主体获得重要奖励及荣誉情况": "entity_honors", "新型经营主体已获得发展配套支持": "supporting_policies",
}


def normalize_header(value: Any) -> str:
    return re.sub(r"\s+", "", clean_text(value) or "").replace("(", "（").replace(")", "）")


@dataclass
class ParsedRow:
    row_number: int
    raw: dict[str, Any]
    basic: dict[str, Any] = field(default_factory=dict)
    education: dict[str, Any] = field(default_factory=dict)
    entity: dict[str, Any] = field(default_factory=dict)
    honors: list[dict[str, Any]] = field(default_factory=list)
    revenues: list[dict[str, Any]] = field(default_factory=list)
    industries: list[dict[str, Any]] = field(default_factory=list)
    cultivation: dict[str, Any] = field(default_factory=dict)


def _grouped(raw: dict[str, Any], prefixes: dict[str, str]) -> list[dict[str, Any]]:
    groups: dict[int, dict[str, Any]] = {}
    for header, value in raw.items():
        for prefix, field_name in prefixes.items():
            match = re.fullmatch(re.escape(prefix) + r"(\d+)?", header)
            if match:
                index = int(match.group(1) or 1)
                groups.setdefault(index, {})[field_name] = value
    return [dict(values, source_column_group=str(index)) for index, values in sorted(groups.items()) if any(v not in (None, "") for v in values.values())]


def parse_workbook(path: Path) -> tuple[str, list[ParsedRow], list[str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Sheet1"] if "Sheet1" in workbook.sheetnames else max(workbook.worksheets, key=lambda ws: ws.max_row * ws.max_column)
    rows = sheet.iter_rows(values_only=True)
    headers = [normalize_header(v) for v in next(rows)]
    unknown = [h for h in headers if h and h not in BASIC_FIELDS and h not in EDUCATION_FIELDS and h not in ENTITY_FIELDS and not re.search(r"\d+$", h) and h not in {"培育年份", "培育诉求", "主营产业", "近三年经营总收入（万元）", "经营年限"}]
    parsed: list[ParsedRow] = []
    for excel_row, values in enumerate(rows, start=2):
        raw = {headers[i]: values[i] for i in range(min(len(headers), len(values))) if headers[i]}
        if not any(v not in (None, "") for v in raw.values()):
            continue
        item = ParsedRow(row_number=excel_row, raw=raw)
        item.basic = {field: raw.get(header) for header, field in BASIC_FIELDS.items() if header in raw}
        item.education = {field: raw.get(header) for header, field in EDUCATION_FIELDS.items() if header in raw}
        item.entity = {field: raw.get(header) for header, field in ENTITY_FIELDS.items() if header in raw}
        item.honors = _grouped(raw, {"荣誉编号": "honor_number", "个人获得荣誉时间": "honor_time", "个人获得荣誉级别": "honor_level", "个人获得荣誉情况": "honor_description"})
        item.revenues = _grouped(raw, {"营收年份": "year", "营业收入（万元）": "operating_revenue", "净利润（万元）": "net_profit", "固定资产净值（万元）": "fixed_asset_net_value", "总资产（万元）": "total_assets", "负债总额（万元）": "total_liabilities", "从业人数（人）": "employee_count", "流动资产（万元）": "current_assets", "管理费用（万元）": "management_expense", "政府补贴金额（万元）": "government_subsidy_amount"})
        item.industries = _grouped(raw, {"主营产业": "industry_name", "近三年经营总收入（万元）": "three_year_total_income", "经营年限": "operation_years"})
        item.cultivation = {"cultivation_year": raw.get("培育年份"), "cultivation_needs": raw.get("培育诉求"), "training_experience": raw.get("培训经历")}
        parsed.append(item)
    workbook.close()
    return sheet.title, parsed, unknown
